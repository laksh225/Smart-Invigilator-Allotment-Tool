#Structure of Excel File
#Index | Name | Email-id | Designation

from openpyxl import load_workbook
from faculty import Faculty
def load():
    wb = load_workbook(filename = "Book1.xlsx")
    sheet1 = wb['Sheet1']

    f_list = []
    p_list = []
    asop_list = []
    assp_list = []
    others = []
    for i in range(1, 100): #i -> row
        if(sheet1.cell(row=i, column=1).value != None):
            #print(str(sheet1.cell(row=i, column=2).value))
            if(str(sheet1.cell(row=i, column=4).value)=='Professor'):
                p_list.append(Faculty(int(sheet1.cell(row=i, column=1).value), str(sheet1.cell(row=i, column=2).value), 0, str(sheet1.cell(row=i, column=4).value)))
            elif(str(sheet1.cell(row=i, column=4).value)=='Associate Professor'): 
                asop_list.append(Faculty(int(sheet1.cell(row=i, column=1).value), str(sheet1.cell(row=i, column=2).value), 0, str(sheet1.cell(row=i, column=4).value)))
            elif(str(sheet1.cell(row=i, column=4).value)=='Assistant Professor'):
                assp_list.append(Faculty(int(sheet1.cell(row=i, column=1).value), str(sheet1.cell(row=i, column=2).value), 0, str(sheet1.cell(row=i, column=4).value)))
            else:
                others.append(Faculty(int(sheet1.cell(row=i, column=1).value), str(sheet1.cell(row=i, column=2).value), 0, str(sheet1.cell(row=i, column=4).value)))
    f_list = [p_list]+[asop_list]+[assp_list]
    assert(len(others)==0)
    return f_list
    
if __name__ == '__main__':
    f_list = load()
    print(*f_list, sep="\n")		# * is print the list without brackets ( '[', ']' )